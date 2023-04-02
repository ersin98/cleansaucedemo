from time import sleep
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import datetime
from datetime import date
import openpyxl 
from globalConstants import GlobalConstants as GC
from selenium.webdriver.support.ui import WebDriverWait as pagewait
from selenium.webdriver.support.expected_conditions import staleness_of

#,from selenium.common.exceptions import NoSuchElementException
class Test_Saucedemo:
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        now = datetime.now()
        self.folderPath = str(now.strftime("%d-%b-%Y")) #str(date.today())
        self.testTime=str(now.strftime("%H.%M")) #.%S
        Path(self.folderPath).mkdir(exist_ok=True)
        self.driver.get(GC.mainURL)

    def teardown_method(self):
        self.driver.quit()

    def page_loaded(self):
        old_page = self.driver.find_element_by_tag_name('html')
        yield
        pagewait(self.driver, 30).until(staleness_of(old_page))

    def waitForElementVisible(self,locator,timeout=10):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

    def getData(excelFileName,sheetName):
        excelFile = openpyxl.load_workbook("data/"+excelFileName)
        selectedSheet = excelFile[sheetName]
        #excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        #selectedSheet=excelFile[sheetName]

        header = {}
        for column in range(1, selectedSheet.max_column + 1):
            header[selectedSheet.cell(row=1, column=column).value] = column
        data = []
        for row in range(2, selectedSheet.max_row + 1):
            row_data = []
            for column in header.values():
                cell_value = selectedSheet.cell(row=row, column=column).value or ""
                row_data.append(cell_value)
            data.append(tuple(row_data))
        return data

    def checkoutpage(self):
        self.cartpage()
        self.waitForElementVisible((By.ID, "checkout"))
        self.driver.find_element(By.ID, "checkout").click()

    def send(self,key,objectId):
        self.waitForElementVisible((By.ID,objectId))
        search_box = self.driver.find_element(By.ID,objectId)
        search_box.clear()
        search_box.send_keys(key)

    def loginclick(self):
        self.waitForElementVisible((By.ID,GC.loginButtonId))
        self.driver.find_element(By.ID,GC.loginButtonId).click()

    def errorMessageWeb(self):
        self.waitForElementVisible((By.XPATH,GC.errorMessageContainerXPATH))
        errorMessageContainer= self.driver.find_element(By.XPATH,GC.errorMessageContainerXPATH)
        return errorMessageContainer.text
    
    def standard_login(self):
        self.send("standard_user",GC.userNameId)
        self.send("secret_sauce",GC.passwordId)
        self.driver.find_element(By.ID,"login-button").click()

    def error_login(self):
        self.send("",GC.userNameId)
        self.send("",GC.passwordId)
        self.driver.find_element(By.ID,"login-button").click()

    def menuAction(self,action):
        self.waitForElementVisible((By.ID,GC.menuButton))
        menu= self.driver.find_element(By.ID,GC.menuButton)
        menu.click()
        self.waitForElementVisible((By.ID,action))
        actionLink= self.driver.find_element(By.ID,action)
        actionLink.click()
    def cartpage(self):
        self.standard_login()
        self.waitForElementVisible((By.CLASS_NAME,"shopping_cart_link"))
        shoppingCart= self.driver.find_element(By.CLASS_NAME,"shopping_cart_link")
        shoppingCart.click()

    @pytest.mark.parametrize("username,password,errorMessageExcel", getData("invalid_login.xlsx","SauceDemoLoginErrors"))
    def test_error_login(self,username,password,errorMessageExcel):
        #hesaba yanlış girişler kontrol edielcek
        self.send(username,str(GC.userNameId))
        self.send(password,str(GC.passwordId))
        self.loginclick()
        errorMessageWeb = self.errorMessageWeb()
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-login-error-{username}-{password}.png")
        assert errorMessageExcel == errorMessageWeb

    def test_standardUserInventory(self):
        #envanter açılacak
        self.standard_login()
        self.waitForElementVisible((By.XPATH,GC.inventoryContainerXPATH))
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-standardUserInventory.png")
        assert self.driver.current_url== GC.inventoryURL

    def test_numerOfItems(self):
        # envanterde 6 adet item olacak
        self.standard_login()
        self.waitForElementVisible((By.CLASS_NAME,GC.inventoryItem))
        products= list(self.driver.find_elements(By.CLASS_NAME,GC.inventoryItem))
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-numerOfItems.png")
        assert len(products)==6

    def test_nullIcon(self):
        #giriş hatalı olduğunda X ikonları görünecek
        self.error_login()
        self.waitForElementVisible((By.XPATH,GC.errorMessageContainerXPATH))
        icon1 = self.driver.find_elements(By.CSS_SELECTOR, ".form_group:nth-child(1) > .svg-inline--fa")
        icon2 = self.driver.find_elements(By.CSS_SELECTOR, ".form_group:nth-child(2) > .svg-inline--fa")
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-nullIcon.png")
        assert (len(icon1) & len(icon2)) > 0

    def test_logout_sidebar_link(self):
        # yandaki menü ile hesaptan çıkılacak
        self.standard_login()
        self.menuAction("logout_sidebar_link")
        self.waitForElementVisible((By.ID,GC.userNameId))
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-logout.png")
        assert self.driver.current_url==GC.mainURL

    def test_cart(self):
        #cart sayfası açılacak
        self.cartpage()
        self.waitForElementVisible((By.ID,"cart_contents_container"))
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-cart.png")
        assert self.driver.current_url==GC.saucelabs
        
    @pytest.mark.xfail(strict=True, reason="blocked")
    def test_twitter(self):
        #twitter logosu twitter sayfasını açamayacak
        self.standard_login()
        self.waitForElementVisible((By.CLASS_NAME,"social_twitter"))
        twitter= self.driver.find_element(By.XPATH,"//*[@id=\"page_wrapper\"]/footer/ul/li[1]/a")
        twitter.click()
        self.driver.implicitly_wait(5)
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-twitter.png")
        assert self.driver.current_url=="https://twitter.com/saucelabs"

    #@pytest.mark.skip(reason="bitmedi")
    @pytest.mark.parametrize("firstName,lastName,zipPostalCode,errorMessage",getData("invalid_login.xlsx","CheckoutErrors"))
    def test_checkout__error_login(self,firstName,lastName,zipPostalCode,errorMessage):
        # chekot sayfasında hatalı girişler kontrol edilecek
        self.checkoutpage()
        self.send(firstName,str("first-name"))
        self.send(lastName,str("last-name"))
        self.send(zipPostalCode,str("postal-code"))
        self.driver.find_element(By.ID, "continue").click()
        errortext = self.driver.find_element(By.XPATH, "//div[@id=\'checkout_info_container\']/div/form/div/div[4]/h3").text
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-checkout-error.png")
        assert errortext== errorMessage

    def test_continue_shopping(self):
        #kart sayfasından continue shopping tuşuyla envanter sayfasına gidilecek
        self.cartpage()
        self.waitForElementVisible((By.ID, "continue-shopping"))
        self.driver.find_element(By.ID, "continue-shopping").click()
        self.page_loaded()
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-continue-shopping.png")
        assert self.driver.current_url== GC.inventoryURL

    def test_back_to_products(self):
        #ürünün detay sayfasından tekrar envanter sayfasına gidilecek
        self.standard_login()
        self.page_loaded()
        self.driver.find_element(By.ID,"item_1_img_link").click()
        self.waitForElementVisible((By.ID,"back-to-products"))
        self.driver.find_element(By.ID,"back-to-products").click()
        self.page_loaded()
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-back-to-products.png")
        assert self.driver.current_url== GC.inventoryURL
    
    def test_inventory_sidebar_link(self):
        #kart sayfasından menüdeki inventory sidebar link tuşuyla envanter sayfasına gidilecek
        self.cartpage()
        self.menuAction("inventory_sidebar_link")
        self.page_loaded()
        self.driver.save_screenshot(f"{self.folderPath}/ {self.testTime}-test-sauce-inventory-sidebar-link.png")
        assert self.driver.current_url== GC.inventoryURL
    
    